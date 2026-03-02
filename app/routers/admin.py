import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, EmailStr
from sqlalchemy.orm import Session
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.database import get_db, User, Anketa, AnketaHistory, UnderwritingRule, RiskRule, EditRequest, Role, SystemSettings
from app.auth import require_permission, hash_password, generate_password, get_user_permissions

router = APIRouter(prefix="/api/admin", tags=["admin"])


# ---------- ROLES CRUD ----------

class RoleOut(BaseModel):
    id: int
    name: str
    is_system: bool
    anketa_create: bool
    anketa_edit: bool
    anketa_view_all: bool
    anketa_conclude: bool
    anketa_delete: bool
    user_manage: bool
    analytics_view: bool
    export_excel: bool
    rules_manage: bool


class CreateRoleRequest(BaseModel):
    name: str = Field(min_length=2, max_length=100)
    anketa_create: bool = False
    anketa_edit: bool = False
    anketa_view_all: bool = False
    anketa_conclude: bool = False
    anketa_delete: bool = False
    user_manage: bool = False
    analytics_view: bool = False
    export_excel: bool = False
    rules_manage: bool = False


class UpdateRoleRequest(BaseModel):
    name: Optional[str] = Field(None, min_length=2, max_length=100)
    anketa_create: Optional[bool] = None
    anketa_edit: Optional[bool] = None
    anketa_view_all: Optional[bool] = None
    anketa_conclude: Optional[bool] = None
    anketa_delete: Optional[bool] = None
    user_manage: Optional[bool] = None
    analytics_view: Optional[bool] = None
    export_excel: Optional[bool] = None
    rules_manage: Optional[bool] = None


PERM_FIELDS = [
    "anketa_create", "anketa_edit", "anketa_view_all", "anketa_conclude",
    "anketa_delete", "user_manage", "analytics_view", "export_excel", "rules_manage",
]


def role_to_out(r: Role) -> RoleOut:
    return RoleOut(
        id=r.id, name=r.name, is_system=r.is_system,
        **{f: getattr(r, f) for f in PERM_FIELDS},
    )


@router.get("/roles", response_model=list[RoleOut])
def list_roles(db: Session = Depends(get_db), user: User = Depends(require_permission("user_manage"))):
    roles = db.query(Role).order_by(Role.id).all()
    return [role_to_out(r) for r in roles]


@router.post("/roles", response_model=RoleOut)
def create_role(
    body: CreateRoleRequest,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("user_manage")),
):
    existing = db.query(Role).filter(Role.name == body.name.strip()).first()
    if existing:
        raise HTTPException(status_code=400, detail="Должность с таким названием уже существует")
    role = Role(name=body.name.strip(), is_system=False)
    for f in PERM_FIELDS:
        setattr(role, f, getattr(body, f))
    db.add(role)
    db.commit()
    db.refresh(role)
    return role_to_out(role)


@router.patch("/roles/{role_id}", response_model=RoleOut)
def update_role(
    role_id: int,
    body: UpdateRoleRequest,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("user_manage")),
):
    role = db.query(Role).filter(Role.id == role_id).first()
    if not role:
        raise HTTPException(status_code=404, detail="Должность не найдена")
    if body.name is not None:
        dup = db.query(Role).filter(Role.name == body.name.strip(), Role.id != role_id).first()
        if dup:
            raise HTTPException(status_code=400, detail="Должность с таким названием уже существует")
        role.name = body.name.strip()
    for f in PERM_FIELDS:
        val = getattr(body, f)
        if val is not None:
            setattr(role, f, val)
    db.commit()
    db.refresh(role)
    return role_to_out(role)


@router.delete("/roles/{role_id}")
def delete_role(
    role_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("user_manage")),
):
    role = db.query(Role).filter(Role.id == role_id).first()
    if not role:
        raise HTTPException(status_code=404, detail="Должность не найдена")
    linked = db.query(User).filter(User.role_id == role_id).count()
    if linked > 0:
        raise HTTPException(status_code=400, detail=f"Нельзя удалить — {linked} пользователь(ей) привязано к этой должности")
    db.delete(role)
    db.commit()
    return {"detail": "Должность удалена"}


# ---------- USERS ----------

class CreateUserRequest(BaseModel):
    email: EmailStr
    full_name: str = Field(min_length=2, max_length=150)
    role_id: int


class UpdateUserRequest(BaseModel):
    full_name: Optional[str] = None
    role_id: Optional[int] = None
    password: Optional[str] = Field(None, min_length=6)
    is_active: Optional[bool] = None
    telegram_chat_id: Optional[str] = None


class UserOut(BaseModel):
    id: int
    email: str
    full_name: str
    role: str
    role_id: Optional[int] = None
    role_name: Optional[str] = None
    is_active: bool
    is_superadmin: bool = False
    created_at: Optional[str] = None
    telegram_chat_id: Optional[str] = None


class CreateUserResponse(UserOut):
    generated_password: Optional[str] = None
    email_sent: bool = False


def user_to_out(u: User) -> UserOut:
    role_name = u.position.name if u.position else ("Администратор" if u.role == "admin" else "Инспектор")
    return UserOut(
        id=u.id, email=u.email, full_name=u.full_name,
        role=u.role, role_id=u.role_id, role_name=role_name,
        is_active=u.is_active, is_superadmin=u.is_superadmin or False,
        created_at=u.created_at.isoformat() if u.created_at else None,
        telegram_chat_id=u.telegram_chat_id,
    )


@router.get("/users", response_model=list[UserOut])
def list_users(db: Session = Depends(get_db), admin: User = Depends(require_permission("user_manage"))):
    users = db.query(User).order_by(User.id).all()
    return [user_to_out(u) for u in users]


@router.post("/users", response_model=CreateUserResponse)
def create_user(
    body: CreateUserRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(require_permission("user_manage")),
):
    existing = db.query(User).filter(User.email == body.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Email уже зарегистрирован")

    role_obj = db.query(Role).filter(Role.id == body.role_id).first()
    if not role_obj:
        raise HTTPException(status_code=400, detail="Должность не найдена")

    # Determine legacy role string for backward compat
    legacy_role = "admin" if role_obj.user_manage else "inspector"

    password = generate_password()

    user = User(
        email=body.email,
        full_name=body.full_name,
        password_hash=hash_password(password),
        role=legacy_role,
        role_id=body.role_id,
        is_active=True,
    )
    db.add(user)
    db.commit()
    db.refresh(user)

    out = user_to_out(user)
    return CreateUserResponse(
        **out.model_dump(),
        generated_password=password,
    )


@router.patch("/users/{user_id}", response_model=UserOut)
def update_user(
    user_id: int,
    body: UpdateUserRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(require_permission("user_manage")),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    if user.is_superadmin and admin.id != user.id:
        raise HTTPException(status_code=403, detail="Нельзя изменить суперадмина")

    if body.full_name is not None:
        user.full_name = body.full_name
    if body.role_id is not None:
        role_obj = db.query(Role).filter(Role.id == body.role_id).first()
        if not role_obj:
            raise HTTPException(status_code=400, detail="Должность не найдена")
        user.role_id = body.role_id
        user.role = "admin" if role_obj.user_manage else "inspector"
    if body.password is not None:
        user.password_hash = hash_password(body.password)
    if body.is_active is not None:
        user.is_active = body.is_active
    if body.telegram_chat_id is not None:
        user.telegram_chat_id = body.telegram_chat_id if body.telegram_chat_id.strip() else None

    db.commit()
    db.refresh(user)
    return user_to_out(user)


@router.post("/users/{user_id}/reset-password")
def reset_password(
    user_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_permission("user_manage")),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    password = generate_password()
    user.password_hash = hash_password(password)
    db.commit()

    return {"email": user.email, "generated_password": password}


@router.delete("/users/{user_id}")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_permission("user_manage")),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    if user.is_superadmin:
        raise HTTPException(status_code=403, detail="Нельзя удалить суперадмина")
    if user.id == admin.id:
        raise HTTPException(status_code=400, detail="Нельзя удалить самого себя")

    # Check if user has related records (anketas, history, etc.)
    has_anketas = db.query(Anketa).filter(Anketa.created_by == user_id).first()
    if has_anketas:
        raise HTTPException(
            status_code=409,
            detail="Нельзя удалить — у пользователя есть анкеты. Деактивируйте его вместо удаления."
        )

    db.delete(user)
    db.commit()
    return {"detail": "Пользователь удалён"}


# ---------- UNDERWRITING RULES ----------

class RuleOutModel(BaseModel):
    id: int
    category: str
    rule_key: str
    value: str
    label: str
    value_type: str


class UpdateRuleBodyRequest(BaseModel):
    value: str


@router.get("/rules", response_model=list[RuleOutModel])
def list_rules(db: Session = Depends(get_db), admin: User = Depends(require_permission("rules_manage"))):
    rules = db.query(UnderwritingRule).order_by(UnderwritingRule.category, UnderwritingRule.id).all()
    return [
        RuleOutModel(
            id=r.id, category=r.category, rule_key=r.rule_key,
            value=r.value, label=r.label, value_type=r.value_type,
        )
        for r in rules
    ]


@router.patch("/rules/{rule_id}", response_model=RuleOutModel)
def update_rule(
    rule_id: int,
    body: UpdateRuleBodyRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(require_permission("rules_manage")),
):
    rule = db.query(UnderwritingRule).filter(UnderwritingRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Правило не найдено")

    val = body.value.strip()
    if rule.value_type == "float":
        try:
            float(val)
        except ValueError:
            raise HTTPException(status_code=400, detail="Значение должно быть числом (float)")
    elif rule.value_type == "int":
        try:
            int(val)
        except ValueError:
            raise HTTPException(status_code=400, detail="Значение должно быть целым числом")
    elif rule.value_type == "string":
        allowed = {"approved", "review", "rejected"}
        if val not in allowed:
            raise HTTPException(status_code=400, detail=f"Допустимые значения: {', '.join(sorted(allowed))}")

    rule.value = val
    db.commit()
    db.refresh(rule)
    return RuleOutModel(
        id=rule.id, category=rule.category, rule_key=rule.rule_key,
        value=rule.value, label=rule.label, value_type=rule.value_type,
    )


# ---------- RISK RULES ----------

class RiskRuleOut(BaseModel):
    id: int
    category: str
    min_pv: float
    is_active: bool


class CreateRiskRuleRequest(BaseModel):
    category: str
    min_pv: float = Field(ge=0)


class UpdateRiskRuleRequest(BaseModel):
    min_pv: Optional[float] = Field(None, ge=0)
    is_active: Optional[bool] = None


@router.get("/risk-rules", response_model=list[RiskRuleOut])
def list_risk_rules(db: Session = Depends(get_db), admin: User = Depends(require_permission("rules_manage"))):
    rules = db.query(RiskRule).order_by(RiskRule.category).all()
    return [
        RiskRuleOut(id=r.id, category=r.category, min_pv=r.min_pv, is_active=r.is_active)
        for r in rules
    ]


@router.post("/risk-rules", response_model=RiskRuleOut)
def create_risk_rule(
    body: CreateRiskRuleRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(require_permission("rules_manage")),
):
    category = body.category.strip()
    if not category:
        raise HTTPException(status_code=400, detail="Категория не может быть пустой")
    existing = db.query(RiskRule).filter(
        RiskRule.category.ilike(category)
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Категория '{category}' уже существует")
    rule = RiskRule(category=category, min_pv=body.min_pv, is_active=True)
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return RiskRuleOut(id=rule.id, category=rule.category, min_pv=rule.min_pv, is_active=rule.is_active)


@router.patch("/risk-rules/{rule_id}", response_model=RiskRuleOut)
def update_risk_rule(
    rule_id: int,
    body: UpdateRiskRuleRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(require_permission("rules_manage")),
):
    rule = db.query(RiskRule).filter(RiskRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Правило не найдено")
    if body.min_pv is not None:
        rule.min_pv = body.min_pv
    if body.is_active is not None:
        rule.is_active = body.is_active
    db.commit()
    db.refresh(rule)
    return RiskRuleOut(id=rule.id, category=rule.category, min_pv=rule.min_pv, is_active=rule.is_active)


@router.delete("/risk-rules/{rule_id}")
def delete_risk_rule(
    rule_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_permission("rules_manage")),
):
    rule = db.query(RiskRule).filter(RiskRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Правило не найдено")
    db.delete(rule)
    db.commit()
    return {"detail": "Правило удалено"}


# ---------- EDIT REQUESTS ----------

class ReviewEditRequest(BaseModel):
    status: str  # approved | rejected
    comment: Optional[str] = None


@router.patch("/edit-requests/{request_id}")
def review_edit_request(
    request_id: int,
    body: ReviewEditRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(require_permission("user_manage")),
):
    """Approve or reject an edit request."""
    req = db.query(EditRequest).filter(EditRequest.id == request_id).first()
    if not req:
        raise HTTPException(status_code=404, detail="Запрос не найден")
    if req.status != "pending":
        raise HTTPException(status_code=400, detail="Запрос уже обработан")
    if body.status not in ("approved", "rejected"):
        raise HTTPException(status_code=400, detail="Статус должен быть «одобрен» или «отклонён»")

    if body.status == "approved":
        anketa = db.query(Anketa).filter(Anketa.id == req.anketa_id).first()
        if anketa:
            old_status = anketa.status
            # Record history
            entry = AnketaHistory(
                anketa_id=anketa.id,
                field_name="status",
                old_value=old_status,
                new_value="draft",
                changed_by=admin.id,
            )
            db.add(entry)
            anketa.status = "draft"

    req.status = body.status
    req.reviewed_by = admin.id
    req.review_comment = body.comment
    req.reviewed_at = datetime.utcnow()

    from app.routers.anketa import create_notification
    status_label = "одобрен" if body.status == "approved" else "отклонён"
    create_notification(db, req.requested_by, "edit_request_reviewed",
        f"Запрос на правку {status_label}",
        f"Ваш запрос на правку анкеты #{req.anketa_id} {status_label}." +
        (f" Комментарий: {body.comment}" if body.comment else ""), req.anketa_id)

    # Telegram notification to requester
    from app.telegram_service import notify_telegram
    notify_telegram(db, req.requested_by,
        f"Запрос на правку анкеты #{req.anketa_id} {status_label}." +
        (f"\nКомментарий: {body.comment}" if body.comment else ""))

    db.commit()
    return {"ok": True}


@router.get("/edit-requests/count")
def get_pending_edit_requests_count(
    db: Session = Depends(get_db),
    admin: User = Depends(require_permission("user_manage")),
):
    """Get count of pending edit requests."""
    count = db.query(EditRequest).filter(EditRequest.status == "pending").count()
    return {"count": count}


# ---------- TELEGRAM SETTINGS ----------

class TelegramSettingsOut(BaseModel):
    bot_token: str | None = None


class TelegramSettingsUpdate(BaseModel):
    bot_token: str | None = None


@router.get("/settings/telegram", response_model=TelegramSettingsOut)
def get_telegram_settings(
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("user_manage")),
):
    setting = db.query(SystemSettings).filter(SystemSettings.key == "telegram_bot_token").first()
    return TelegramSettingsOut(bot_token=setting.value if setting else None)


@router.patch("/settings/telegram", response_model=TelegramSettingsOut)
def update_telegram_settings(
    body: TelegramSettingsUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("user_manage")),
):
    setting = db.query(SystemSettings).filter(SystemSettings.key == "telegram_bot_token").first()
    token_val = body.bot_token.strip() if body.bot_token else None
    if setting:
        setting.value = token_val
    else:
        db.add(SystemSettings(key="telegram_bot_token", value=token_val))
    db.commit()
    return TelegramSettingsOut(bot_token=token_val)


# ---------- EXCEL EXPORT ----------

def _fmt(val):
    """Format value for Excel cell."""
    if val is None:
        return ""
    return val


def _fmt_date(val):
    """Format date/datetime for Excel."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d.%m.%Y %H:%M")
    return str(val)


def _fmt_num(val):
    """Format number for Excel."""
    if val is None:
        return ""
    return val


def _status_label(status):
    labels = {
        "draft": "Черновик",
        "saved": "Сохранена",
        "approved": "Одобрена",
        "review": "На рассмотрении",
        "rejected_underwriter": "Отказ андеррайтера",
        "rejected_client": "Отказ клиента",
        "deleted": "Удалена",
    }
    return labels.get(status, status or "")


def _decision_label(decision):
    labels = {
        "approved": "Одобрено",
        "review": "На рассмотрении",
        "rejected_underwriter": "Отказ андеррайтера",
        "rejected_client": "Отказ клиента",
    }
    return labels.get(decision, decision or "")


def _write_header(ws, headers):
    """Write styled header row."""
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="6B3FA0", end_color="6B3FA0", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="E4DDEF"),
        right=Side(style="thin", color="E4DDEF"),
        bottom=Side(style="thin", color="E4DDEF"),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.freeze_panes = "A2"


def _fill_individual_sheet(ws, anketas, users_map):
    """Fill sheet with individual anketa data."""
    headers = [
        "ID", "Дата создания", "Статус", "ФИО клиента", "Дата рождения",
        "ПИНФЛ", "Паспорт", "Телефон", "Адрес фактический",
        "Партнёр", "Марка", "Модель", "Год авто", "Пробег (км)",
        "Стоимость (сум)", "ПВ %", "ПВ сумма", "Остаток",
        "Срок (мес)", "Ставка %", "Ежемесячный платёж",
        "Доход", "DTI %", "Категория просрочки",
        "Решение", "Авто-вердикт", "Рекоменд. ПВ %",
        "Комментарий", "Андеррайтер", "Дата заключения",
    ]
    _write_header(ws, headers)

    for row, a in enumerate(anketas, 2):
        concluder = users_map.get(a.concluded_by, "") if a.concluded_by else ""
        ws.cell(row=row, column=1, value=a.id)
        ws.cell(row=row, column=2, value=_fmt_date(a.created_at))
        ws.cell(row=row, column=3, value=_status_label(a.status))
        ws.cell(row=row, column=4, value=_fmt(a.full_name))
        ws.cell(row=row, column=5, value=_fmt_date(a.birth_date) if a.birth_date else "")
        ws.cell(row=row, column=6, value=_fmt(a.pinfl))
        ws.cell(row=row, column=7, value=_fmt(a.passport_series))
        ws.cell(row=row, column=8, value=_fmt(a.phone_numbers))
        ws.cell(row=row, column=9, value=_fmt(a.actual_address))
        ws.cell(row=row, column=10, value=_fmt(a.partner))
        ws.cell(row=row, column=11, value=_fmt(a.car_brand))
        ws.cell(row=row, column=12, value=_fmt(a.car_model))
        ws.cell(row=row, column=13, value=_fmt_num(a.car_year))
        ws.cell(row=row, column=14, value=_fmt_num(a.mileage))
        ws.cell(row=row, column=15, value=_fmt_num(a.purchase_price))
        ws.cell(row=row, column=16, value=_fmt_num(a.down_payment_percent))
        ws.cell(row=row, column=17, value=_fmt_num(a.down_payment_amount))
        ws.cell(row=row, column=18, value=_fmt_num(a.remaining_amount))
        ws.cell(row=row, column=19, value=_fmt_num(a.lease_term_months))
        ws.cell(row=row, column=20, value=_fmt_num(a.interest_rate))
        ws.cell(row=row, column=21, value=_fmt_num(a.monthly_payment))
        ws.cell(row=row, column=22, value=_fmt_num(a.total_monthly_income))
        ws.cell(row=row, column=23, value=_fmt_num(a.dti))
        ws.cell(row=row, column=24, value=_fmt(a.overdue_category))
        ws.cell(row=row, column=25, value=_decision_label(a.decision))
        ws.cell(row=row, column=26, value=_fmt(a.auto_decision))
        ws.cell(row=row, column=27, value=_fmt_num(a.recommended_pv))
        ws.cell(row=row, column=28, value=_fmt(a.conclusion_comment))
        ws.cell(row=row, column=29, value=concluder)
        ws.cell(row=row, column=30, value=_fmt_date(a.concluded_at))

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 16


def _fill_legal_entity_sheet(ws, anketas, users_map):
    """Fill sheet with legal entity anketa data."""
    headers = [
        "ID", "Дата создания", "Статус",
        "Наименование компании", "ИНН", "ОКЭД",
        "Юр. адрес", "Факт. адрес", "Телефон компании",
        "ФИО директора", "Телефон директора",
        "Контактное лицо", "Должность", "Телефон конт. лица",
        "Марка", "Модель", "Год авто", "Пробег (км)",
        "Стоимость (сум)", "ПВ %", "ПВ сумма", "Остаток",
        "Срок (мес)", "Ставка %", "Ежемесячный платёж",
        "Выручка компании", "Чистая прибыль", "Доход директора",
        "Общий месячный доход", "DTI %",
        "Просрочка компании", "Просрочка директора", "Просрочка поручителя",
        "Поручитель ФИО", "Поручитель ПИНФЛ", "Поручитель доход",
        "Решение", "Авто-вердикт", "Рекоменд. ПВ %",
        "Комментарий", "Андеррайтер", "Дата заключения",
    ]
    _write_header(ws, headers)

    for row, a in enumerate(anketas, 2):
        concluder = users_map.get(a.concluded_by, "") if a.concluded_by else ""
        ws.cell(row=row, column=1, value=a.id)
        ws.cell(row=row, column=2, value=_fmt_date(a.created_at))
        ws.cell(row=row, column=3, value=_status_label(a.status))
        ws.cell(row=row, column=4, value=_fmt(a.company_name))
        ws.cell(row=row, column=5, value=_fmt(a.company_inn))
        ws.cell(row=row, column=6, value=_fmt(a.company_oked))
        ws.cell(row=row, column=7, value=_fmt(a.company_legal_address))
        ws.cell(row=row, column=8, value=_fmt(a.company_actual_address))
        ws.cell(row=row, column=9, value=_fmt(a.company_phone))
        ws.cell(row=row, column=10, value=_fmt(a.director_full_name))
        ws.cell(row=row, column=11, value=_fmt(a.director_phone))
        ws.cell(row=row, column=12, value=_fmt(a.contact_person_name))
        ws.cell(row=row, column=13, value=_fmt(a.contact_person_role))
        ws.cell(row=row, column=14, value=_fmt(a.contact_person_phone))
        ws.cell(row=row, column=15, value=_fmt(a.car_brand))
        ws.cell(row=row, column=16, value=_fmt(a.car_model))
        ws.cell(row=row, column=17, value=_fmt_num(a.car_year))
        ws.cell(row=row, column=18, value=_fmt_num(a.mileage))
        ws.cell(row=row, column=19, value=_fmt_num(a.purchase_price))
        ws.cell(row=row, column=20, value=_fmt_num(a.down_payment_percent))
        ws.cell(row=row, column=21, value=_fmt_num(a.down_payment_amount))
        ws.cell(row=row, column=22, value=_fmt_num(a.remaining_amount))
        ws.cell(row=row, column=23, value=_fmt_num(a.lease_term_months))
        ws.cell(row=row, column=24, value=_fmt_num(a.interest_rate))
        ws.cell(row=row, column=25, value=_fmt_num(a.monthly_payment))
        ws.cell(row=row, column=26, value=_fmt_num(a.company_revenue_total))
        ws.cell(row=row, column=27, value=_fmt_num(a.company_net_profit))
        ws.cell(row=row, column=28, value=_fmt_num(a.director_income_total))
        ws.cell(row=row, column=29, value=_fmt_num(a.total_monthly_income))
        ws.cell(row=row, column=30, value=_fmt_num(a.dti))
        ws.cell(row=row, column=31, value=_fmt(a.company_overdue_category))
        ws.cell(row=row, column=32, value=_fmt(a.director_overdue_category))
        ws.cell(row=row, column=33, value=_fmt(a.guarantor_overdue_category))
        ws.cell(row=row, column=34, value=_fmt(a.guarantor_full_name))
        ws.cell(row=row, column=35, value=_fmt(a.guarantor_pinfl))
        ws.cell(row=row, column=36, value=_fmt_num(a.guarantor_monthly_income))
        ws.cell(row=row, column=37, value=_decision_label(a.decision))
        ws.cell(row=row, column=38, value=_fmt(a.auto_decision))
        ws.cell(row=row, column=39, value=_fmt_num(a.recommended_pv))
        ws.cell(row=row, column=40, value=_fmt(a.conclusion_comment))
        ws.cell(row=row, column=41, value=concluder)
        ws.cell(row=row, column=42, value=_fmt_date(a.concluded_at))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 16


@router.get("/export-excel")
def export_excel(
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    db: Session = Depends(get_db),
    admin: User = Depends(require_permission("export_excel")),
):
    """Export anketas to Excel with separate sheets for individuals and legal entities."""
    query = db.query(Anketa)

    if date_from:
        try:
            dt_from = datetime.fromisoformat(date_from)
            query = query.filter(Anketa.created_at >= dt_from)
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный формат date_from")

    if date_to:
        try:
            from datetime import timedelta
            dt_to = datetime.fromisoformat(date_to) + timedelta(days=1)
            query = query.filter(Anketa.created_at <= dt_to)
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный формат date_to")

    anketas = query.order_by(Anketa.id.desc()).all()

    # Build users map for concluder names
    users = db.query(User).all()
    users_map = {u.id: u.full_name for u in users}

    individuals = [a for a in anketas if (getattr(a, 'client_type', None) or "individual") == "individual"]
    legal_entities = [a for a in anketas if getattr(a, 'client_type', None) == "legal_entity"]

    wb = Workbook()

    # Sheet 1: Individuals
    ws_ind = wb.active
    ws_ind.title = "Физические лица"
    _fill_individual_sheet(ws_ind, individuals, users_map)

    # Sheet 2: Legal entities
    ws_leg = wb.create_sheet("Юридические лица")
    _fill_legal_entity_sheet(ws_leg, legal_entities, users_map)

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"anketas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
